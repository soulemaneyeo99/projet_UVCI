from fastapi import APIRouter, Depends, HTTPException
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session
from typing import Optional
import io

from app.db.database import get_db
from app.models.models import Teacher, Activity, Resource, Course, AcademicYear, QuotaStatutaire
from app.core.security import require_authenticated, require_admin_or_secretary

router = APIRouter()


@router.get("/pdf/{teacher_id}")
def export_teacher_pdf(
    teacher_id: int,
    db: Session = Depends(get_db),
    _=Depends(require_authenticated),
):
    try:
        from reportlab.lib.pagesizes import A4
        from reportlab.lib import colors
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
        from reportlab.lib.units import cm
    except ImportError:
        raise HTTPException(status_code=500, detail="ReportLab non disponible. Installez reportlab.")

    teacher = db.query(Teacher).filter(Teacher.id == teacher_id).first()
    if not teacher:
        raise HTTPException(status_code=404, detail="Enseignant non trouvé")

    activities = db.query(Activity).filter(Activity.teacher_id == teacher_id).all()

    buffer = io.BytesIO()
    doc = SimpleDocTemplate(
        buffer,
        pagesize=A4,
        leftMargin=2 * cm,
        rightMargin=2 * cm,
        topMargin=2 * cm,
        bottomMargin=2 * cm,
    )

    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        "CustomTitle",
        parent=styles["Heading1"],
        fontSize=16,
        textColor=colors.HexColor("#1F4E79"),
        spaceAfter=12,
    )
    subtitle_style = ParagraphStyle(
        "CustomSubtitle",
        parent=styles["Heading2"],
        fontSize=13,
        textColor=colors.HexColor("#2E75B6"),
        spaceAfter=8,
    )

    story = []
    story.append(Paragraph("UNIVERSITÉ VIRTUELLE DE CÔTE D'IVOIRE", title_style))
    story.append(Paragraph("Récapitulatif des Heures Pédagogiques", subtitle_style))
    story.append(Spacer(1, 0.5 * cm))

    # Teacher information table
    info_data = [
        ["Nom et Prénom :", f"{teacher.prenom} {teacher.nom}"],
        ["Grade :", teacher.grade],
        ["Département :", teacher.departement],
        ["Statut :", teacher.statut],
        ["Email :", teacher.email],
        ["Téléphone :", teacher.telephone or "-"],
    ]
    info_table = Table(info_data, colWidths=[5 * cm, 12 * cm])
    info_table.setStyle(
        TableStyle([
            ("FONTNAME", (0, 0), (-1, -1), "Helvetica"),
            ("FONTSIZE", (0, 0), (-1, -1), 10),
            ("FONTNAME", (0, 0), (0, -1), "Helvetica-Bold"),
            ("BACKGROUND", (0, 0), (0, -1), colors.HexColor("#EBF2F7")),
            ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
            ("ROWBACKGROUNDS", (0, 0), (-1, -1), [colors.white, colors.HexColor("#F5F7FA")]),
            ("TOPPADDING", (0, 0), (-1, -1), 4),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
        ])
    )
    story.append(info_table)
    story.append(Spacer(1, 0.7 * cm))

    # Activities detail table
    story.append(Paragraph("Détail des Activités", subtitle_style))

    act_data = [["Cours", "Type", "Niveau", "Séquences", "Volume (h)", "Statut"]]
    total_volume = 0.0

    for act in activities:
        resource = db.query(Resource).filter(Resource.id == act.resource_id).first()
        course_name = "N/A"
        niveau = 1
        if resource:
            niveau = resource.niveau_complexite
            course = db.query(Course).filter(Course.id == resource.course_id).first()
            if course:
                course_name = course.intitule[:35]

        type_label = "Création" if act.type == "creation" else "Mise à jour"
        status_label = "Validé" if act.validation_status == "valide" else "En attente"

        act_data.append([
            course_name,
            type_label,
            f"Niveau {niveau}",
            str(act.nb_sequences),
            f"{act.volume_horaire_calcule:.2f}",
            status_label,
        ])
        
        if act.validation_status == "valide":
            total_volume += act.volume_horaire_calcule

    act_data.append(["", "", "", "TOTAL VALIDÉ", f"{total_volume:.2f} h", ""])

    act_table = Table(
        act_data,
        colWidths=[5.5 * cm, 3 * cm, 2.5 * cm, 2.5 * cm, 2.5 * cm, 2.5 * cm],
    )
    act_table.setStyle(
        TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1F4E79")),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
            ("FONTSIZE", (0, 0), (-1, -1), 9),
            ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
            ("ROWBACKGROUNDS", (0, 1), (-1, -2), [colors.white, colors.HexColor("#F5F7FA")]),
            ("BACKGROUND", (0, -1), (-1, -1), colors.HexColor("#EBF2F7")),
            ("FONTNAME", (0, -1), (-1, -1), "Helvetica-Bold"),
            ("ALIGN", (3, 0), (-1, -1), "CENTER"),
            ("TOPPADDING", (0, 0), (-1, -1), 4),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
        ])
    )
    story.append(act_table)
    story.append(Spacer(1, 0.5 * cm))

    # Fetch dynamic quota
    quota = db.query(QuotaStatutaire).filter(
        QuotaStatutaire.grade == teacher.grade,
        QuotaStatutaire.statut == teacher.statut
    ).first()
    seuil_horaire = quota.quota_heures if quota else 192.0

    # Summary line
    summary_data = [
        ["Volume horaire validé :", f"{total_volume:.2f} h"],
        [f"Heures complémentaires (au-delà de {seuil_horaire}h) :", f"{max(0.0, total_volume - seuil_horaire):.2f} h"],
        ["Montant estimé (FCFA) :", f"{max(0.0, total_volume - seuil_horaire) * teacher.taux_horaire:,.0f}"],
    ]
    summary_table = Table(summary_data, colWidths=[9 * cm, 5 * cm])
    summary_table.setStyle(
        TableStyle([
            ("FONTNAME", (0, 0), (-1, -1), "Helvetica"),
            ("FONTSIZE", (0, 0), (-1, -1), 10),
            ("FONTNAME", (0, 0), (0, -1), "Helvetica-Bold"),
            ("ALIGN", (1, 0), (1, -1), "RIGHT"),
            ("TOPPADDING", (0, 0), (-1, -1), 3),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 3),
        ])
    )
    story.append(summary_table)

    doc.build(story)
    buffer.seek(0)

    filename = f"recapitulatif_{teacher.nom}_{teacher.prenom}.pdf"
    return StreamingResponse(
        buffer,
        media_type="application/pdf",
        headers={"Content-Disposition": f"attachment; filename={filename}"},
    )


@router.get("/excel")
def export_global_excel(
    departement: Optional[str] = None,
    db: Session = Depends(get_db),
    _=Depends(require_admin_or_secretary),
):
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
    except ImportError:
        raise HTTPException(status_code=500, detail="openpyxl non disponible. Installez openpyxl.")

    wb = Workbook()
    ws = wb.active
    ws.title = "État des Heures"

    # Style helpers
    header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True, size=11)
    center = Alignment(horizontal="center", vertical="center")
    right = Alignment(horizontal="right", vertical="center")
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )
    alt_fill = PatternFill(start_color="EBF2F7", end_color="EBF2F7", fill_type="solid")

    # Title row
    ws.merge_cells("A1:H1")
    ws["A1"] = "UNIVERSITÉ VIRTUELLE DE CÔTE D'IVOIRE – État Global des Heures Pédagogiques"
    ws["A1"].font = Font(bold=True, size=13, color="1F4E79")
    ws["A1"].alignment = center
    ws.row_dimensions[1].height = 28

    # Blank separator
    ws.row_dimensions[2].height = 8

    # Column headers (row 3)
    headers = [
        "Nom", "Prénom", "Grade", "Département", "Statut",
        "Heures Totales", "Heures Complémentaires", "Montant Estimé (FCFA)",
    ]
    col_widths = [18, 18, 22, 22, 14, 16, 22, 22]
    for col, (header, width) in enumerate(zip(headers, col_widths), 1):
        cell = ws.cell(row=3, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center
        cell.border = thin_border
        ws.column_dimensions[get_column_letter(col)].width = width
    ws.row_dimensions[3].height = 20

    # Data rows
    query = db.query(Teacher)
    if departement:
        query = query.filter(Teacher.departement == departement)
    teachers = query.all()

    for row_num, teacher in enumerate(teachers, 4):
        acts = db.query(Activity).filter(
            Activity.teacher_id == teacher.id,
            Activity.validation_status == 'valide'
        ).all()
        volume = sum(a.volume_horaire_calcule for a in acts)
        
        # Dynamic quota
        quota = db.query(QuotaStatutaire).filter(
            QuotaStatutaire.grade == teacher.grade,
            QuotaStatutaire.statut == teacher.statut
        ).first()
        seuil = quota.quota_heures if quota else 192.0

        heures_comp = max(0.0, volume - seuil)
        montant = heures_comp * teacher.taux_horaire

        row_data = [
            teacher.nom,
            teacher.prenom,
            teacher.grade,
            teacher.departement,
            teacher.statut,
            round(volume, 2),
            round(heures_comp, 2),
            round(montant, 0),
        ]
        use_alt = row_num % 2 == 0
        for col, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_num, column=col, value=value)
            cell.border = thin_border
            if use_alt:
                cell.fill = alt_fill
            if col >= 6:
                cell.alignment = right

    # Total row
    last_data_row = len(teachers) + 3
    total_row = last_data_row + 1
    ws.cell(row=total_row, column=1, value="TOTAL").font = Font(bold=True)
    if last_data_row >= 4:
        ws.cell(
            row=total_row, column=6,
            value=f"=SUM(F4:F{last_data_row})"
        ).font = Font(bold=True)
        ws.cell(
            row=total_row, column=7,
            value=f"=SUM(G4:G{last_data_row})"
        ).font = Font(bold=True)
        ws.cell(
            row=total_row, column=8,
            value=f"=SUM(H4:H{last_data_row})"
        ).font = Font(bold=True)

    # Freeze header rows
    ws.freeze_panes = "A4"

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    return StreamingResponse(
        buffer,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=etat_heures_UVCI.xlsx"},
    )
